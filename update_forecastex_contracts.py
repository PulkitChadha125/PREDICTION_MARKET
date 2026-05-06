"""Update pairs/contracts from ForecastEx and append new symbols to Excel.

Workflow:
1) Open https://forecastex.com/data and find the latest date row.
2) Download latest "Pairs" CSV.
3) Extract symbols from `event_contract` and compare with previous CSV symbols.
4) For new symbols, call local PredictionMarket API (/events/topics/console).
5) Append lookup rows to prediction_market_symbols_from_pairs.xlsx.
6) Optionally clean up temporary downloaded files.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from openpyxl import Workbook, load_workbook

from ibkrsearch import extract_symbol, fetch_symbol_topics, response_preview

DATA_PAGE_URL = "https://forecastex.com/data"
DEFAULT_MAIN_XLSX = "prediction_market_symbols_from_pairs.xlsx"
DEFAULT_PREVIOUS_CSV = "pairs_20260414.csv"
HEADERS = [
    "generated_at_utc",
    "source_csv",
    "csv_row_number",
    "event_contract",
    "lookup_symbol",
    "index",
    "symbol",
    "name",
    "conid",
    "exchange",
    "months",
    "note",
    "response_preview",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download latest ForecastEx pairs CSV and append new symbols to Excel."
    )
    parser.add_argument("--data-url", default=DATA_PAGE_URL, help="ForecastEx data page URL.")
    parser.add_argument(
        "--previous-csv",
        default=DEFAULT_PREVIOUS_CSV,
        help="Existing pairs CSV used as baseline for symbol comparison.",
    )
    parser.add_argument(
        "--downloads-dir",
        default="downloads",
        help="Directory to save downloaded latest pairs file.",
    )
    parser.add_argument(
        "--main-xlsx",
        default=DEFAULT_MAIN_XLSX,
        help="Main Excel file to append new symbol lookups into.",
    )
    parser.add_argument(
        "--base-url",
        default="http://127.0.0.1:8000",
        help="PredictionMarket API base URL.",
    )
    parser.add_argument(
        "--exchange",
        default="FORECASTX",
        help="Exchange passed to /events/topics/console.",
    )
    parser.add_argument("--timeout", type=float, default=30.0, help="HTTP timeout seconds.")
    parser.add_argument(
        "--print-new-symbols",
        action="store_true",
        help="Print full list of newly discovered symbols before IBKR fetch starts.",
    )
    parser.add_argument(
        "--replace-previous-csv",
        action="store_true",
        help="Replace previous CSV file with latest downloaded CSV after processing.",
    )
    parser.add_argument(
        "--cleanup-download",
        action="store_true",
        help="Delete downloaded latest CSV after processing.",
    )
    return parser.parse_args()


def _strip_tags(text: str) -> str:
    return re.sub(r"<[^>]+>", "", text)


def find_latest_pairs_download(data_url: str, timeout: float) -> tuple[str, str]:
    """Return (latest_date, pairs_download_url)."""
    resp = requests.get(data_url, timeout=timeout)
    resp.raise_for_status()
    html = resp.text

    rows = re.findall(r"<tr[^>]*>(.*?)</tr>", html, flags=re.IGNORECASE | re.DOTALL)
    latest_date = ""
    latest_pairs_url = ""

    for row_html in rows:
        clean = _strip_tags(row_html)
        date_match = re.search(r"\b(\d{4}-\d{2}-\d{2})\b", clean)
        if not date_match:
            continue

        row_date = date_match.group(1)
        hrefs = re.findall(
            r"""<a[^>]+href=["']([^"']+)["'][^>]*>""",
            row_html,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if not hrefs:
            continue

        # In ForecastEx table order: Pairs, Prices, Summary -> use first href.
        pairs_href = hrefs[0]
        latest_date = row_date
        latest_pairs_url = urljoin(data_url, pairs_href)
        break

    if not latest_date or not latest_pairs_url:
        raise RuntimeError("Could not find latest pairs download link on ForecastEx data page.")
    return latest_date, latest_pairs_url


def download_file(url: str, output_path: Path, timeout: float) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with requests.get(url, timeout=timeout, stream=True) as resp:
        resp.raise_for_status()
        with output_path.open("wb") as handle:
            for chunk in resp.iter_content(chunk_size=1024 * 64):
                if chunk:
                    handle.write(chunk)


def symbols_from_pairs_csv(csv_path: Path) -> set[str]:
    symbols: set[str] = set()
    with csv_path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        if "event_contract" not in (reader.fieldnames or []):
            raise ValueError(f"`event_contract` column missing in {csv_path}")
        for row in reader:
            symbol = extract_symbol(str(row.get("event_contract") or ""))
            if symbol:
                symbols.add(symbol)
    return symbols


def ensure_workbook(path: Path) -> tuple[Any, Any]:
    if path.exists():
        wb = load_workbook(path)
        sheet = wb.active
        # If workbook exists but header is missing, rewrite header on row 1.
        if sheet.max_row == 0:
            sheet.append(HEADERS)
        elif sheet.cell(row=1, column=1).value != HEADERS[0]:
            sheet.insert_rows(1)
            for idx, header in enumerate(HEADERS, start=1):
                sheet.cell(row=1, column=idx, value=header)
        return wb, sheet

    wb = Workbook()
    sheet = wb.active
    sheet.title = "topics_from_csv"
    sheet.append(HEADERS)
    return wb, sheet


def existing_lookup_keys(sheet: Any) -> set[tuple[str, str]]:
    """Build key set (lookup_symbol, conid_str) to avoid duplicate appends."""
    keys: set[tuple[str, str]] = set()
    # Header columns based on ibkrsearch.py schema:
    # E=lookup_symbol, I=conid
    for row_idx in range(2, sheet.max_row + 1):
        symbol = sheet.cell(row=row_idx, column=5).value
        conid = sheet.cell(row=row_idx, column=9).value
        if symbol is None:
            continue
        keys.add((str(symbol), "" if conid is None else str(conid)))
    return keys


def append_new_symbol_rows(
    sheet: Any,
    latest_csv_path: Path,
    new_symbols: list[str],
    base_url: str,
    exchange: str,
    timeout: float,
) -> tuple[int, int]:
    generated_at = datetime.now(UTC).isoformat()
    appended_rows = 0
    matched_rows = 0
    existing_keys = existing_lookup_keys(sheet)

    for index, symbol in enumerate(new_symbols, start=1):
        if index % 50 == 1:
            print(f"[{index}] fetching symbol={symbol}")
        payload, error = fetch_symbol_topics(base_url, symbol, exchange, timeout)

        if error:
            key = (symbol, "")
            if key in existing_keys:
                continue
            row = [
                generated_at,
                str(latest_csv_path),
                None,
                None,
                symbol,
                None,
                None,
                None,
                None,
                exchange,
                None,
                None,
                error,
            ]
            sheet.append(row)
            existing_keys.add(key)
            appended_rows += 1
            continue

        topics = payload.get("topics", []) if isinstance(payload, dict) else []
        note = payload.get("note") if isinstance(payload, dict) else None

        if not isinstance(topics, list) or not topics:
            key = (symbol, "")
            if key in existing_keys:
                continue
            row = [
                generated_at,
                str(latest_csv_path),
                None,
                None,
                symbol,
                None,
                None,
                None,
                None,
                exchange,
                None,
                note,
                response_preview(payload),
            ]
            sheet.append(row)
            existing_keys.add(key)
            appended_rows += 1
            continue

        for topic in topics:
            if not isinstance(topic, dict):
                continue
            conid = topic.get("conid")
            key = (symbol, "" if conid is None else str(conid))
            if key in existing_keys:
                continue

            months = topic.get("months")
            months_value = ";".join(str(m) for m in months) if isinstance(months, list) else ""
            row = [
                generated_at,
                str(latest_csv_path),
                None,
                None,
                symbol,
                topic.get("index"),
                topic.get("symbol"),
                topic.get("name"),
                conid,
                topic.get("exchange"),
                months_value,
                note,
                response_preview(payload),
            ]
            sheet.append(row)
            existing_keys.add(key)
            appended_rows += 1
            if conid is not None:
                matched_rows += 1

    return appended_rows, matched_rows


def main() -> None:
    args = parse_args()
    previous_csv = Path(args.previous_csv)
    downloads_dir = Path(args.downloads_dir)
    main_xlsx = Path(args.main_xlsx)

    if not previous_csv.exists():
        raise FileNotFoundError(f"Previous CSV not found: {previous_csv}")

    latest_date, latest_pairs_url = find_latest_pairs_download(args.data_url, args.timeout)
    latest_csv_path = downloads_dir / f"pairs_{latest_date.replace('-', '')}.csv"
    print(f"Latest date: {latest_date}")
    print(f"Downloading pairs: {latest_pairs_url}")
    download_file(latest_pairs_url, latest_csv_path, args.timeout)

    previous_symbols = symbols_from_pairs_csv(previous_csv)
    latest_symbols = symbols_from_pairs_csv(latest_csv_path)
    new_symbols = sorted(latest_symbols - previous_symbols)

    print(f"Previous symbols: {len(previous_symbols)}")
    print(f"Latest symbols:   {len(latest_symbols)}")
    print(f"New symbols:      {len(new_symbols)}")
    if args.print_new_symbols and new_symbols:
        print("New symbols to fetch from IBKR:")
        print(", ".join(new_symbols))

    if new_symbols:
        wb, sheet = ensure_workbook(main_xlsx)
        appended_rows, matched_rows = append_new_symbol_rows(
            sheet=sheet,
            latest_csv_path=latest_csv_path,
            new_symbols=new_symbols,
            base_url=args.base_url,
            exchange=args.exchange,
            timeout=args.timeout,
        )
        wb.save(main_xlsx)
        print(f"Excel updated: {main_xlsx}")
        print(f"Rows appended: {appended_rows} | matched topic rows: {matched_rows}")
    else:
        print("No new symbols found. Excel not modified.")

    if args.replace_previous_csv:
        previous_csv.write_bytes(latest_csv_path.read_bytes())
        print(f"Replaced baseline CSV: {previous_csv}")

    if args.cleanup_download and latest_csv_path.exists():
        latest_csv_path.unlink()
        print(f"Deleted downloaded file: {latest_csv_path}")

    print("Done.")


if __name__ == "__main__":
    main()
