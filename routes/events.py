"""Event/contract discovery routes."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from fastapi import APIRouter, HTTPException, Query

from models.event_models import (
    AllTopicsResponse,
    EventChainRequest,
    ConsoleTopicItem,
    ConsoleTopicListResponse,
    ContractInfoResponse,
    EventChainResponse,
    EventTopicSearchResponse,
    MarketQuoteItem,
    MarketQuoteResponse,
    StrikeListResponse,
)
from services.event_service import EventService
from services.ibkr_client import IBKRClient, IBKRClientError

router = APIRouter(prefix="/events", tags=["events"])
event_service = EventService(IBKRClient())
PAIRS_SYMBOLS_XLSX = (
    Path(__file__).resolve().parent.parent / "prediction_market_symbols_from_pairs.xlsx"
)


def _sec_type_candidates(sec_type: str) -> list[str]:
    """Rank secType attempts for discovery fallback."""
    requested = sec_type.upper()
    candidates = [requested]
    for fallback in ("FOP", "FUT", "IND", "OPT"):
        if fallback not in candidates:
            candidates.append(fallback)
    return candidates


def _exchange_aliases(exchange_upper: str) -> set[str]:
    """Normalize exchange aliases used by IBKR for CME-group products."""
    aliases = {exchange_upper}
    if exchange_upper == "CME":
        aliases.update({"CBT", "CBOT", "XCBT", "XCME", "COMEX", "NYMEX"})
    elif exchange_upper in {"CBT", "CBOT"}:
        aliases.update({"CBT", "CBOT", "XCBT", "CME"})
    elif exchange_upper == "NYMEX":
        aliases.update({"NYMEX", "XNYM", "CME"})
    elif exchange_upper == "COMEX":
        aliases.update({"COMEX", "XCEC", "CME"})
    return aliases


def _topic_matches_exchange(topic: object, exchange_upper: str) -> bool:
    """Match topic against exchange using raw exchange + description + sections."""
    raw = getattr(topic, "raw", {}) or {}
    topic_desc = str(getattr(topic, "description", "") or raw.get("description", "")).upper()
    topic_exchange = str(getattr(topic, "exchange", "") or raw.get("exchange", "")).upper()
    raw_sections = raw.get("sections", [])
    if not isinstance(raw_sections, list):
        raw_sections = []

    section_exchanges = {
        str(section.get("exchange", "")).upper()
        for section in raw_sections
        if isinstance(section, dict)
    }
    has_ec_section = any(
        isinstance(section, dict) and str(section.get("secType", "")).upper() == "EC"
        for section in raw_sections
    )

    aliases = _exchange_aliases(exchange_upper)
    match_in_text = any(alias in topic_desc or alias in topic_exchange for alias in aliases)
    match_in_sections = bool(section_exchanges.intersection(aliases))

    if exchange_upper == "FORECASTX":
        return match_in_text or match_in_sections or has_ec_section
    if match_in_text or match_in_sections:
        return True
    # If no exchange metadata at all, keep row (IBKR can return sparse secdef records).
    return not (topic_exchange or section_exchanges)


def _build_chain_response(
    *,
    symbol: str | None,
    conid: str | None,
    sec_type: str,
    month: str,
    exchange: str,
    sectype: str,
) -> EventChainResponse:
    """Shared chain builder for query and body-based APIs."""
    if conid:
        contracts = event_service.build_chain_from_conid(
            conid=conid, month=month, exchange=exchange, sectype=sectype
        )
        topic_symbol = symbol or f"CONID:{conid}"
    else:
        if not symbol:
            raise HTTPException(
                status_code=422,
                detail="Provide either symbol or conid for chain lookup.",
            )
        contracts = event_service.build_chain(
            symbol=symbol,
            sec_type=sec_type,
            month=month,
            exchange=exchange,
            sectype=sectype,
        )
        topic_symbol = symbol
    return EventChainResponse(
        status="success",
        topic_symbol=topic_symbol,
        month=month,
        exchange=exchange,
        contracts=contracts,
    )


def _to_optional_float(value: object) -> float | None:
    """Convert IBKR snapshot numeric field to float, handling N/A-like values."""
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip().upper()
        if not cleaned or cleaned in {"N/A", "-"}:
            return None
        value = cleaned
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _quote_field_count(row: dict[str, object]) -> int:
    """Count populated quote fields among LTP/BID/ASK/VOLUME for row quality ranking."""
    populated = 0
    for field in ("31", "84", "86", "87"):
        if _to_optional_float(row.get(field)) is not None:
            populated += 1
    return populated


def _pick_richer_row(
    current: dict[str, object] | None, candidate: dict[str, object]
) -> dict[str, object]:
    """Pick snapshot row with richer quote payload for same conid."""
    if current is None:
        return candidate

    current_score = (_quote_field_count(current), len(current))
    candidate_score = (_quote_field_count(candidate), len(candidate))
    if candidate_score > current_score:
        return candidate
    return current


@router.get("/search", response_model=EventTopicSearchResponse)
def search_events(
    symbol: str = Query(..., description="Underlying symbol, e.g. NQ"),
    sec_type: str = Query("IND", description="Security type for search"),
) -> EventTopicSearchResponse:
    """Search event topics/underlyings using IBKR secdef search."""
    try:
        topics = event_service.search_topics(symbol=symbol, sec_type=sec_type)
        return EventTopicSearchResponse(
            status="success",
            symbol=symbol,
            sec_type=sec_type,
            topics=topics,
        )
    except IBKRClientError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/topics/all", response_model=AllTopicsResponse)
@router.get("/symbols", response_model=AllTopicsResponse)
def get_all_topics(
    exchange: str | None = Query(
        "FORECASTX",
        description="Optional exchange filter, default FORECASTX for prediction markets.",
    ),
) -> AllTopicsResponse:
    """
    Fetch all prediction-market **product** symbols (underlying/topic roots) from IBKR.

    IBKR Campus Event Trading documents ForecastEx as OPT on exchange FORECASTX and
    notes that market scanners are not available for event contracts, so there is no
    generic \"list every tradable YES/NO line\" discovery call at the API level.

    For a **full catalog of product roots** (symbol, conid, name, category), IBKR
    exposes the Client Portal **category tree** (`/trsrv/event/category-tree`), which
    this endpoint consumes via ``IBKRClient.get_event_category_tree``.

    To expand one product into strikes and YES/NO contract conids for a month, use
    ``GET /events/chain`` (secdef strikes + info), which follows the options-style
    workflow described in the Event Trading guide.

    Same handler as ``/events/topics/all``; ``/events/symbols`` is a clearer alias.

    Example:
    curl -k "http://127.0.0.1:8000/events/symbols?exchange=FORECASTX"
    curl -k "http://127.0.0.1:8000/events/topics/all?exchange=FORECASTX"
    """
    try:
        topics = event_service.get_all_prediction_topics(exchange_filter=exchange)
        return AllTopicsResponse(status="success", total_topics=len(topics), topics=topics)
    except IBKRClientError as exc:
        if "category tree endpoint is not available" in str(exc).lower():
            return AllTopicsResponse(
                status="success",
                total_topics=0,
                topics=[],
                note=str(exc),
            )
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/topics/console", response_model=ConsoleTopicListResponse)
def get_prediction_topics_console(
    symbols: str | None = Query(
        None,
        description=(
            "Comma-separated topic symbols to probe, e.g. FF,USIP,CPI,GDP. "
            "If omitted, API will try to fetch all topics from category-tree endpoint."
        ),
    ),
    exchange: str | None = Query(
        "FORECASTX",
        description="Optional exchange filter, e.g. FORECASTX, CME, CBT.",
    ),
    sec_type: str = Query(
        "IND",
        description="secdef search type, e.g. IND (ForecastEx), FOP/FUT (CME-style).",
    ),
) -> ConsoleTopicListResponse:
    """
    Return compact prediction-market topic list for console usage.

    Example:
    curl -k "http://127.0.0.1:8000/events/topics/console?symbols=FF,USIP&exchange=FORECASTX"
    """
    try:
        exchange_upper = exchange.upper() if exchange else None
        sec_type_upper = sec_type.upper()
        # If symbols are not provided, attempt full category-tree retrieval.
        if not symbols:
            try:
                all_topics = event_service.get_all_prediction_topics(
                    exchange_filter=exchange_upper
                )
                compact_rows = [
                    ConsoleTopicItem(
                        index=idx,
                        symbol=t.symbol,
                        name=t.name,
                        conid=t.conid,
                        exchange=t.exchange,
                        months=[],
                    )
                    for idx, t in enumerate(all_topics, start=1)
                ]
                return ConsoleTopicListResponse(
                    status="success",
                    total_topics=len(compact_rows),
                    topics=compact_rows,
                )
            except IBKRClientError as exc:
                return ConsoleTopicListResponse(
                    status="success",
                    total_topics=0,
                    topics=[],
                    note=(
                        "All-topics endpoint is unavailable on this gateway build. "
                        "Pass symbols list explicitly, e.g. ?symbols=FF,USIP."
                        f" Details: {exc}"
                    ),
                )

        cleaned_symbols = [s.strip().upper() for s in symbols.split(",") if s.strip()]
        seen: set[int] = set()
        compact_rows: list[ConsoleTopicItem] = []

        for sym in cleaned_symbols:
            topics = []
            for candidate_sec_type in _sec_type_candidates(sec_type_upper):
                topics = event_service.search_topics(symbol=sym, sec_type=candidate_sec_type)
                if topics:
                    break
            for topic in topics:
                if topic.conid is None:
                    continue

                raw = topic.raw or {}
                if exchange_upper and not _topic_matches_exchange(topic, exchange_upper):
                    continue
                if topic.conid in seen:
                    continue

                compact_rows.append(
                    ConsoleTopicItem(
                        index=0,  # assigned below after sorting
                        symbol=str(topic.symbol or sym),
                        name=str(raw.get("companyName") or topic.symbol or sym),
                        conid=int(topic.conid),
                        exchange=exchange_upper,
                        months=topic.months,
                    )
                )
                seen.add(topic.conid)

        compact_rows.sort(key=lambda row: (row.symbol, row.name, row.conid))
        for idx, row in enumerate(compact_rows, start=1):
            row.index = idx

        return ConsoleTopicListResponse(
            status="success",
            total_topics=len(compact_rows),
            topics=compact_rows,
        )
    except IBKRClientError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/strikes", response_model=StrikeListResponse)
def get_strikes(
    conid: str,
    sectype: str,
    month: str,
    exchange: str,
) -> StrikeListResponse:
    """Return available strikes for selected contract month."""
    try:
        strikes, broker_response = event_service.get_strikes(
            conid=conid, sectype=sectype, month=month, exchange=exchange
        )
        return StrikeListResponse(
            conid=conid,
            sectype=sectype,
            month=month,
            exchange=exchange,
            strikes=strikes,
            broker_response=broker_response,
        )
    except IBKRClientError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/pairs/symbols", tags=["events"])
def get_pairs_symbols_xlsx_json(
    sheet_name: str | None = Query(
        default=None,
        description="Optional sheet name. Defaults to the workbook active sheet.",
    ),
) -> dict[str, object]:
    """Return `prediction_market_symbols_from_pairs.xlsx` rows as JSON."""
    if not PAIRS_SYMBOLS_XLSX.exists():
        raise HTTPException(
            status_code=404,
            detail=f"File not found: {PAIRS_SYMBOLS_XLSX.name}",
        )

    try:
        workbook = load_workbook(PAIRS_SYMBOLS_XLSX, data_only=True, read_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
    except KeyError as exc:
        available = [name for name in workbook.sheetnames]
        raise HTTPException(
            status_code=400,
            detail=f"Sheet '{sheet_name}' not found. Available sheets: {available}",
        ) from exc
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"Failed to read xlsx: {exc}") from exc

    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return {
            "status": "success",
            "file": PAIRS_SYMBOLS_XLSX.name,
            "sheet": sheet.title,
            "total_rows": 0,
            "rows": [],
        }

    headers = [str(value).strip() if value is not None else "" for value in header_row]
    data_rows: list[dict[str, object | None]] = []

    for row in rows_iter:
        if row is None:
            continue
        record: dict[str, object | None] = {}
        is_empty = True
        for idx, cell_value in enumerate(row):
            header = headers[idx] if idx < len(headers) and headers[idx] else f"column_{idx + 1}"
            value: object | None = cell_value
            if isinstance(value, datetime):
                value = value.isoformat()
            elif isinstance(value, date):
                value = value.isoformat()
            if value not in (None, ""):
                is_empty = False
            record[header] = value
        if not is_empty:
            data_rows.append(record)

    return {
        "status": "success",
        "file": PAIRS_SYMBOLS_XLSX.name,
        "sheet": sheet.title,
        "total_rows": len(data_rows),
        "rows": data_rows,
    }


@router.get("/contracts", response_model=ContractInfoResponse)
def get_contracts(
    conid: str,
    sectype: str,
    month: str,
    exchange: str,
    strike: float,
) -> ContractInfoResponse:
    """Get raw contract info for a single strike."""
    try:
        contracts = event_service.get_contracts_for_strike(
            conid=conid,
            sectype=sectype,
            month=month,
            exchange=exchange,
            strike=strike,
        )
        return ContractInfoResponse(
            conid=conid,
            sectype=sectype,
            month=month,
            exchange=exchange,
            strike=strike,
            contracts=contracts,
        )
    except IBKRClientError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/chain", response_model=EventChainResponse)
def get_chain(
    symbol: str | None = Query(
        default=None,
        description="Optional underlying symbol, e.g. NQ. Required when conid is not provided.",
    ),
    sec_type: str = Query("IND"),
    month: str = Query(..., description="Option month, e.g. 202605"),
    exchange: str = Query(..., description="Exchange code"),
    sectype: str = Query("OPT", description="Contract sec type used for strikes/info"),
    conid: str | None = Query(
        default=None,
        description="Optional known underlying conid to skip search and speed up chain fetch.",
    ),
) -> EventChainResponse:
    """
    Build normalized YES/NO chain.

    Usage in Swagger/docs:
    - Fast path: provide ``conid`` + ``month`` + ``exchange`` (+ optional ``symbol`` label).
    - Discovery path: provide ``symbol`` + ``month`` + ``exchange`` and omit ``conid``.

    Example:
    curl -k "http://127.0.0.1:8000/events/chain?symbol=NQ&sec_type=IND&month=202605&exchange=CME&sectype=OPT"
    """
    try:
        return _build_chain_response(
            symbol=symbol,
            conid=conid,
            sec_type=sec_type,
            month=month,
            exchange=exchange,
            sectype=sectype,
        )
    except IBKRClientError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.post("/chain/check", response_model=EventChainResponse)
def check_chain(body: EventChainRequest) -> EventChainResponse:
    """
    Swagger-friendly chain check using request body values.

    Use this when you want to enter your own symbol/conid/month/exchange directly
    in one payload and verify contracts quickly from `/docs`.
    """
    try:
        return _build_chain_response(
            symbol=body.symbol,
            conid=body.conid,
            sec_type=body.sec_type,
            month=body.month,
            exchange=body.exchange,
            sectype=body.sectype,
        )
    except IBKRClientError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/discover", response_model=EventTopicSearchResponse)
def discover_underlyings(
    symbol: str = Query(..., description="Event product symbol, e.g. NQ or FF"),
    sec_type: str = Query("IND", description="Usually IND for event underlyings"),
    exchange: str | None = Query(
        default=None, description="Optional exchange filter, e.g. CME or FORECASTX"
    ),
) -> EventTopicSearchResponse:
    """
    Discover underlying records and conids before chain retrieval.

    Example:
    curl -k "http://127.0.0.1:8000/events/discover?symbol=NQ&sec_type=IND&exchange=CME"
    """
    try:
        topics = []
        seen_conids: set[int] = set()
        for candidate_sec_type in _sec_type_candidates(sec_type):
            candidate_topics = event_service.search_topics(
                symbol=symbol, sec_type=candidate_sec_type
            )
            for topic in candidate_topics:
                if topic.conid is None:
                    topics.append(topic)
                    continue
                if topic.conid in seen_conids:
                    continue
                seen_conids.add(topic.conid)
                topics.append(topic)
            if topics:
                break
        if exchange:
            exchange_upper = exchange.upper()
            topics = [t for t in topics if _topic_matches_exchange(t, exchange_upper)]
        return EventTopicSearchResponse(
            status="success", symbol=symbol, sec_type=sec_type, topics=topics
        )
    except IBKRClientError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get(
    "/quotes",
    response_model=MarketQuoteResponse,
    summary="Get LTP/BID/ASK/VOLUME snapshot",
    description=(
        "Fetch quote snapshot for one or multiple contract conids. "
        "By default returns IBKR fields 31(LTP), 84(BID), 86(ASK), and 87(VOLUME)."
    ),
)
def get_quotes(
    conids: str = Query(
        ...,
        description="Single conid or comma-separated conids, e.g. 877309547,877309550",
        examples=["877309547,877309550"],
    ),
    fields: str = Query(
        "31,84,86,87",
        description="IBKR marketdata fields. Defaults to 31(LTP),84(BID),86(ASK),87(VOLUME).",
        examples=["31,84,86,87"],
    ),
) -> MarketQuoteResponse:
    """Fetch quote snapshot (LTP/BID/ASK/VOLUME) for one or multiple contract conids."""
    parsed_conids = [chunk.strip() for chunk in conids.split(",") if chunk.strip()]
    if not parsed_conids:
        raise HTTPException(status_code=422, detail="Provide at least one conid.")

    unique_conids = list(dict.fromkeys(parsed_conids))
    parsed_conid_ints: list[int] = []
    try:
        parsed_conid_ints = [int(conid) for conid in unique_conids]
        snapshot_rows = event_service.client.get_market_snapshot(
            conids=parsed_conid_ints,
            fields=fields,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail="All conids must be integers.") from exc
    except IBKRClientError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc

    rows_by_conid: dict[int, dict[str, object]] = {}
    for row in snapshot_rows:
        if not isinstance(row, dict):
            continue
        conid_val = row.get("conid")
        try:
            conid_int = int(conid_val)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            continue
        rows_by_conid[conid_int] = _pick_richer_row(rows_by_conid.get(conid_int), row)

    # IBKR snapshot is often sparse on first fetch; retry missing rows in batch, then one-by-one.
    def _missing_quote(conid_int: int) -> bool:
        row = rows_by_conid.get(conid_int, {})
        return (
            _to_optional_float(row.get("31")) is None
            and _to_optional_float(row.get("84")) is None
            and _to_optional_float(row.get("86")) is None
            and _to_optional_float(row.get("87")) is None
        )

    missing_conids = [cid for cid in parsed_conid_ints if _missing_quote(cid)]
    if missing_conids:
        try:
            retry_rows = event_service.client.get_market_snapshot(
                conids=missing_conids,
                fields=fields,
            )
            for row in retry_rows:
                if not isinstance(row, dict):
                    continue
                conid_val = row.get("conid")
                try:
                    conid_int = int(conid_val)  # type: ignore[arg-type]
                except (TypeError, ValueError):
                    continue
                rows_by_conid[conid_int] = _pick_richer_row(
                    rows_by_conid.get(conid_int), row
                )
        except IBKRClientError:
            # Keep best-effort data from initial snapshot.
            pass

    missing_conids = [cid for cid in parsed_conid_ints if _missing_quote(cid)]
    for conid_int in missing_conids:
        try:
            single_rows = event_service.client.get_market_snapshot(
                conids=[conid_int],
                fields=fields,
            )
        except IBKRClientError:
            continue
        for row in single_rows:
            if not isinstance(row, dict):
                continue
            conid_val = row.get("conid")
            try:
                row_conid = int(conid_val)  # type: ignore[arg-type]
            except (TypeError, ValueError):
                continue
            rows_by_conid[row_conid] = _pick_richer_row(rows_by_conid.get(row_conid), row)

    quotes: list[MarketQuoteItem] = []
    for conid in unique_conids:
        conid_int = int(conid)
        raw = rows_by_conid.get(conid_int, {})
        quotes.append(
            MarketQuoteItem(
                conid=conid_int,
                ltp=_to_optional_float(raw.get("31")),
                bid=_to_optional_float(raw.get("84")),
                ask=_to_optional_float(raw.get("86")),
                volume=_to_optional_float(raw.get("87")),
                raw=raw,
            )
        )

    return MarketQuoteResponse(status="success", total=len(quotes), quotes=quotes)
