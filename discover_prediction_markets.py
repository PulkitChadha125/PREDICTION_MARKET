"""Probe prediction market endpoints with useful parameter permutations.

Run:
    python discover_prediction_markets.py

Optional:
    python discover_prediction_markets.py --base-url http://127.0.0.1:8000 --max-chain-topics 50
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import requests

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover - optional dependency in some runtimes
    Workbook = None
    load_workbook = None


@dataclass
class FetchRecord:
    """Human-readable record of one API fetch."""

    title: str
    endpoint: str
    params: dict[str, Any]
    ok: bool
    status_code: int
    elapsed_ms: int
    summary: str
    response_preview: Any
    error: str | None = None


class DiscoveryRunner:
    """Executes permutations and keeps a structured report."""

    def __init__(
        self,
        base_url: str,
        timeout: int,
        symbol_limit: int,
        chunk_size: int,
        max_chain_topics: int,
    ) -> None:
        self.base_url = base_url.rstrip("/")
        self.timeout = timeout
        self.symbol_limit = symbol_limit
        self.chunk_size = chunk_size
        self.max_chain_topics = max_chain_topics
        self.session = requests.Session()
        self.records: list[FetchRecord] = []
        self.unique_topics_by_conid: dict[int, dict[str, Any]] = {}
        self.chain_probe_results: list[dict[str, Any]] = []
        self.yes_no_contract_details: list[dict[str, Any]] = []

    def _url(self, endpoint: str) -> str:
        return f"{self.base_url}{endpoint}"

    def fetch(
        self,
        title: str,
        endpoint: str,
        params: dict[str, Any] | None = None,
    ) -> dict[str, Any] | list[Any] | None:
        """Call one endpoint, print details, and save a report row."""
        params = params or {}
        url = self._url(endpoint)
        print("\n" + "=" * 88)
        print(f"[FETCH] {title}")
        print(f"  URL: {url}")
        if params:
            print(f"  Params: {json.dumps(params, ensure_ascii=True)}")

        try:
            started = datetime.now(tz=UTC)
            response = self.session.get(url, params=params, timeout=self.timeout)
            elapsed_ms = int((datetime.now(tz=UTC) - started).total_seconds() * 1000)
            status_code = response.status_code

            try:
                payload = response.json()
            except ValueError:
                payload = {"non_json_response": response.text[:500]}

            if response.ok:
                summary = self._summarize_payload(payload)
                print(f"  Status: {status_code} OK ({elapsed_ms} ms)")
                print(f"  Summary: {summary}")
                preview = self._preview(payload)
                if preview is not None:
                    print("  Preview:")
                    print(json.dumps(preview, indent=2, ensure_ascii=True))
                self.records.append(
                    FetchRecord(
                        title=title,
                        endpoint=endpoint,
                        params=params,
                        ok=True,
                        status_code=status_code,
                        elapsed_ms=elapsed_ms,
                        summary=summary,
                        response_preview=preview,
                    )
                )
                return payload

            err_text = payload.get("detail") if isinstance(payload, dict) else str(payload)
            print(f"  Status: {status_code} ERROR ({elapsed_ms} ms)")
            print(f"  Error: {err_text}")
            self.records.append(
                FetchRecord(
                    title=title,
                    endpoint=endpoint,
                    params=params,
                    ok=False,
                    status_code=status_code,
                    elapsed_ms=elapsed_ms,
                    summary="request_failed",
                    response_preview=self._preview(payload),
                    error=str(err_text),
                )
            )
        except requests.RequestException as exc:
            print(f"  Request failed: {exc}")
            self.records.append(
                FetchRecord(
                    title=title,
                    endpoint=endpoint,
                    params=params,
                    ok=False,
                    status_code=0,
                    elapsed_ms=0,
                    summary="network_error",
                    response_preview=None,
                    error=str(exc),
                )
            )
        return None

    @staticmethod
    def _preview(payload: Any) -> Any:
        if isinstance(payload, dict):
            if "topics" in payload and isinstance(payload["topics"], list):
                return {
                    "status": payload.get("status"),
                    "total_topics": payload.get("total_topics"),
                    "topics_sample": payload["topics"][:3],
                    "note": payload.get("note"),
                }
            if "contracts" in payload and isinstance(payload["contracts"], list):
                return {"contracts_count": len(payload["contracts"]), "sample": payload["contracts"][:2]}
            return payload
        if isinstance(payload, list):
            return payload[:3]
        return payload

    @staticmethod
    def _summarize_payload(payload: Any) -> str:
        if isinstance(payload, dict):
            if "topics" in payload and isinstance(payload["topics"], list):
                return f"topics={len(payload['topics'])}"
            if "contracts" in payload and isinstance(payload["contracts"], list):
                return f"contracts={len(payload['contracts'])}"
            if "authenticated" in payload:
                return (
                    f"authenticated={payload.get('authenticated')}, "
                    f"brokerage_ready={payload.get('brokerage_ready', 'n/a')}"
                )
            return f"keys={','.join(sorted(payload.keys()))[:120]}"
        if isinstance(payload, list):
            return f"list_len={len(payload)}"
        return str(type(payload))

    def ingest_topics(self, payload: dict[str, Any] | None, source: str) -> int:
        """Add unique topics by conid from one endpoint result."""
        if not isinstance(payload, dict):
            return 0
        rows = payload.get("topics")
        if not isinstance(rows, list):
            return 0

        added = 0
        for row in rows:
            if not isinstance(row, dict):
                continue
            conid = row.get("conid")
            if not isinstance(conid, int):
                continue
            if conid in self.unique_topics_by_conid:
                continue

            self.unique_topics_by_conid[conid] = {
                "conid": conid,
                "symbol": row.get("symbol"),
                "name": row.get("name") or row.get("description"),
                "exchange": row.get("exchange"),
                "months": row.get("months", []),
                "source": source,
                "raw": row,
            }
            added += 1
        return added

    def run_auth_checks(self) -> None:
        self.fetch("Auth status", "/auth/status")
        self.fetch("Auth readiness", "/auth/ready")

    def run_topic_endpoint_permutations(self) -> None:
        """Try exchange permutations for category-tree-backed endpoints."""
        exchange_variants = ["FORECASTX", "CME", "CBOT", "SMART"]

        for exch in exchange_variants:
            data = self.fetch(
                f"Topics/all exchange={exch}",
                "/events/topics/all",
                {"exchange": exch},
            )
            added = self.ingest_topics(data if isinstance(data, dict) else None, source=f"topics_all:{exch}")
            if added:
                print(f"  Added unique topics: {added}")

        # Also call without exchange filter if backend allows it.
        data = self.fetch("Topics/all exchange omitted", "/events/topics/all")
        added = self.ingest_topics(data if isinstance(data, dict) else None, source="topics_all:none")
        if added:
            print(f"  Added unique topics: {added}")

    def run_console_permutations(self, symbols: list[str]) -> None:
        """Try /events/topics/console with and without symbol chunks."""
        data = self.fetch(
            "Topics/console all (exchange=FORECASTX)",
            "/events/topics/console",
            {"exchange": "FORECASTX"},
        )
        added = self.ingest_topics(data if isinstance(data, dict) else None, source="topics_console:all")
        if added:
            print(f"  Added unique topics: {added}")

        for idx, chunk in enumerate(chunks(symbols, self.chunk_size), start=1):
            symbols_str = ",".join(chunk)
            data = self.fetch(
                f"Topics/console chunk {idx}",
                "/events/topics/console",
                {"symbols": symbols_str, "exchange": "FORECASTX"},
            )
            added = self.ingest_topics(
                data if isinstance(data, dict) else None,
                source=f"topics_console:chunk_{idx}",
            )
            print(f"  Chunk {idx} symbols={len(chunk)} -> new unique topics={added}")

    def run_search_permutations(self, symbols: list[str]) -> None:
        """Run symbol x sec_type combinations for /events/search."""
        sec_types = ["IND", "STK", "FUT", "CMDTY"]

        for symbol in symbols[: self.symbol_limit]:
            for sec_type in sec_types:
                payload = self.fetch(
                    f"Search symbol={symbol} sec_type={sec_type}",
                    "/events/search",
                    {"symbol": symbol, "sec_type": sec_type},
                )
                if not isinstance(payload, dict):
                    continue

                rows = payload.get("topics", [])
                if not isinstance(rows, list):
                    continue

                filtered = []
                for row in rows:
                    if not isinstance(row, dict):
                        continue
                    desc = str(row.get("description") or "").upper()
                    exch = str(row.get("exchange") or "").upper()
                    raw = row.get("raw") if isinstance(row.get("raw"), dict) else {}
                    sections = raw.get("sections") if isinstance(raw, dict) else []
                    has_ec_section = isinstance(sections, list) and any(
                        isinstance(s, dict) and str(s.get("secType", "")).upper() == "EC"
                        for s in sections
                    )
                    if "FORECAST" in desc or "FORECAST" in exch or has_ec_section:
                        filtered.append(row)

                wrapped = {"topics": filtered}
                added = self.ingest_topics(wrapped, source=f"search:{symbol}:{sec_type}")
                if added:
                    print(f"  Search added unique topics={added}")

    def run_chain_permutations(self) -> None:
        """Probe chain endpoint to validate contracts per discovered topic."""
        topics = list(self.unique_topics_by_conid.values())[: self.max_chain_topics]
        if not topics:
            print("\nNo topics available for chain probing.")
            return

        month_candidates = month_variants()
        sectype_candidates = ["OPT", "FOP"]

        print("\n" + "=" * 88)
        print("[CHAIN PROBES] Running symbol x month x sectype checks on discovered topics")

        for topic in topics:
            symbol = str(topic.get("symbol") or "").strip()
            conid = topic.get("conid")
            exchange = str(topic.get("exchange") or "FORECASTX")
            months = normalize_months(topic.get("months"))
            if not months:
                months = month_candidates[:2]
            else:
                months = months[:2]

            for month in months:
                for sectype in sectype_candidates:
                    payload = self.fetch(
                        f"Chain symbol={symbol} conid={conid} month={month} sectype={sectype}",
                        "/events/chain",
                        {
                            "symbol": symbol or "UNKNOWN",
                            "sec_type": "IND",
                            "month": month,
                            "exchange": exchange,
                            "sectype": sectype,
                            "conid": str(conid),
                        },
                    )
                    contracts_count = 0
                    yes_conid_sample: int | None = None
                    no_conid_sample: int | None = None
                    if isinstance(payload, dict) and isinstance(payload.get("contracts"), list):
                        contracts_count = len(payload["contracts"])
                        for row in payload["contracts"]:
                            if not isinstance(row, dict):
                                continue
                            strike_value = row.get("strike")
                            yes_leg = row.get("yes_contract")
                            no_leg = row.get("no_contract")

                            yes_has = isinstance(yes_leg, dict) and isinstance(yes_leg.get("conid"), int)
                            no_has = isinstance(no_leg, dict) and isinstance(no_leg.get("conid"), int)

                            self.yes_no_contract_details.append(
                                {
                                    "symbol": symbol,
                                    "topic_conid": conid,
                                    "exchange": exchange,
                                    "month": month,
                                    "sectype": sectype,
                                    "strike": strike_value,
                                    "has_yes": bool(yes_has),
                                    "has_no": bool(no_has),
                                    "yes_conid": yes_leg.get("conid") if isinstance(yes_leg, dict) else None,
                                    "no_conid": no_leg.get("conid") if isinstance(no_leg, dict) else None,
                                    "yes_right": yes_leg.get("right") if isinstance(yes_leg, dict) else None,
                                    "no_right": no_leg.get("right") if isinstance(no_leg, dict) else None,
                                    "yes_description": (
                                        yes_leg.get("description") if isinstance(yes_leg, dict) else None
                                    ),
                                    "no_description": (
                                        no_leg.get("description") if isinstance(no_leg, dict) else None
                                    ),
                                    "yes_maturity_date": (
                                        yes_leg.get("maturity_date") if isinstance(yes_leg, dict) else None
                                    ),
                                    "no_maturity_date": (
                                        no_leg.get("maturity_date") if isinstance(no_leg, dict) else None
                                    ),
                                    "yes_trading_class": (
                                        yes_leg.get("trading_class") if isinstance(yes_leg, dict) else None
                                    ),
                                    "no_trading_class": (
                                        no_leg.get("trading_class") if isinstance(no_leg, dict) else None
                                    ),
                                }
                            )

                            if isinstance(yes_leg, dict) and isinstance(yes_leg.get("conid"), int):
                                if yes_conid_sample is None:
                                    yes_conid_sample = int(yes_leg["conid"])
                            if isinstance(no_leg, dict) and isinstance(no_leg.get("conid"), int):
                                if no_conid_sample is None:
                                    no_conid_sample = int(no_leg["conid"])
                    self.chain_probe_results.append(
                        {
                            "symbol": symbol,
                            "conid": conid,
                            "exchange": exchange,
                            "month": month,
                            "sectype": sectype,
                            "contracts_count": contracts_count,
                            "yes_conid_sample": yes_conid_sample,
                            "no_conid_sample": no_conid_sample,
                        }
                    )

    def print_final_summary(self) -> None:
        print("\n" + "=" * 88)
        print("FINAL SUMMARY")
        print("=" * 88)
        print(f"Total fetches: {len(self.records)}")
        print(f"Successful fetches: {sum(1 for r in self.records if r.ok)}")
        print(f"Failed fetches: {sum(1 for r in self.records if not r.ok)}")
        print(f"Unique topics found: {len(self.unique_topics_by_conid)}")

        per_symbol: dict[str, int] = {}
        for row in self.unique_topics_by_conid.values():
            sym = str(row.get("symbol") or "UNKNOWN")
            per_symbol[sym] = per_symbol.get(sym, 0) + 1

        top_symbols = sorted(per_symbol.items(), key=lambda kv: (-kv[1], kv[0]))[:30]
        print("\nTop symbols by count:")
        for sym, count in top_symbols:
            print(f"  {sym:<12} {count}")

        if self.chain_probe_results:
            non_empty = [r for r in self.chain_probe_results if r["contracts_count"] > 0]
            print(f"\nChain probes run: {len(self.chain_probe_results)}")
            print(f"Chain probes with contracts > 0: {len(non_empty)}")
            if self.yes_no_contract_details:
                complete_pairs = [
                    r for r in self.yes_no_contract_details if r["has_yes"] and r["has_no"]
                ]
                print(
                    "Strike rows captured: "
                    f"{len(self.yes_no_contract_details)} (complete YES/NO pairs: {len(complete_pairs)})"
                )

    def write_report(self, output_path: Path) -> None:
        output = {
            "generated_at_utc": datetime.now(tz=UTC).isoformat(),
            "base_url": self.base_url,
            "records": [record.__dict__ for record in self.records],
            "unique_topics": sorted(
                self.unique_topics_by_conid.values(),
                key=lambda row: (str(row.get("symbol") or ""), int(row.get("conid") or 0)),
            ),
            "chain_probe_results": self.chain_probe_results,
            "yes_no_contract_details": self.yes_no_contract_details,
        }
        output_path.write_text(json.dumps(output, indent=2, ensure_ascii=True), encoding="utf-8")
        print(f"\nDetailed report written to: {output_path}")

    def _best_chain_by_conid(self) -> dict[int, dict[str, Any]]:
        best: dict[int, dict[str, Any]] = {}
        for row in self.chain_probe_results:
            conid = row.get("conid")
            if not isinstance(conid, int):
                continue
            existing = best.get(conid)
            if existing is None:
                best[conid] = row
                continue

            # Prefer more contracts; tie-break in favor of OPT for event contracts.
            existing_count = int(existing.get("contracts_count") or 0)
            current_count = int(row.get("contracts_count") or 0)
            if current_count > existing_count:
                best[conid] = row
                continue
            if current_count == existing_count:
                existing_is_opt = str(existing.get("sectype") or "").upper() == "OPT"
                current_is_opt = str(row.get("sectype") or "").upper() == "OPT"
                if current_is_opt and not existing_is_opt:
                    best[conid] = row
        return best

    def write_excel(self, output_path: Path) -> None:
        """Write discovered symbols and chain outcomes into an Excel workbook."""
        if Workbook is None:
            print(
                f"\nExcel export skipped because openpyxl is not available. "
                f"Install it to write: {output_path}"
            )
            return

        workbook = Workbook()
        symbols_sheet = workbook.active
        symbols_sheet.title = "prediction_symbols"
        symbols_sheet.append(
            [
                "symbol",
                "conid",
                "name",
                "exchange",
                "months",
                "source",
                "search_url",
                "chain_url_template",
                "best_month",
                "best_sectype",
                "best_contracts_count",
                "yes_conid_sample",
                "no_conid_sample",
            ]
        )

        best_by_conid = self._best_chain_by_conid()
        for row in sorted(
            self.unique_topics_by_conid.values(),
            key=lambda item: (str(item.get("symbol") or ""), int(item.get("conid") or 0)),
        ):
            conid = int(row.get("conid"))
            symbol = str(row.get("symbol") or "")
            exchange = str(row.get("exchange") or "FORECASTX")
            best = best_by_conid.get(conid, {})
            months = normalize_months(row.get("months"))
            month_hint = best.get("month") or (months[0] if months else "")
            symbols_sheet.append(
                [
                    symbol,
                    conid,
                    row.get("name") or "",
                    exchange,
                    ";".join(months),
                    row.get("source") or "",
                    f"{self.base_url}/events/search?symbol={symbol}&sec_type=IND",
                    (
                        f"{self.base_url}/events/chain?symbol={symbol}&sec_type=IND"
                        f"&month={month_hint or 'YYYYMM'}&exchange={exchange}&sectype=OPT&conid={conid}"
                    ),
                    best.get("month") or "",
                    best.get("sectype") or "",
                    int(best.get("contracts_count") or 0),
                    best.get("yes_conid_sample") or "",
                    best.get("no_conid_sample") or "",
                ]
            )

        chain_sheet = workbook.create_sheet("chain_probe_results")
        chain_sheet.append(
            [
                "symbol",
                "conid",
                "exchange",
                "month",
                "sectype",
                "contracts_count",
                "yes_conid_sample",
                "no_conid_sample",
            ]
        )
        for row in self.chain_probe_results:
            chain_sheet.append(
                [
                    row.get("symbol"),
                    row.get("conid"),
                    row.get("exchange"),
                    row.get("month"),
                    row.get("sectype"),
                    row.get("contracts_count"),
                    row.get("yes_conid_sample"),
                    row.get("no_conid_sample"),
                ]
            )

        strikes_sheet = workbook.create_sheet("yes_no_by_strike")
        strikes_sheet.append(
            [
                "symbol",
                "topic_conid",
                "exchange",
                "month",
                "sectype",
                "strike",
                "has_yes",
                "has_no",
                "yes_conid",
                "no_conid",
                "yes_right",
                "no_right",
                "yes_description",
                "no_description",
                "yes_maturity_date",
                "no_maturity_date",
                "yes_trading_class",
                "no_trading_class",
            ]
        )
        for row in self.yes_no_contract_details:
            strikes_sheet.append(
                [
                    row.get("symbol"),
                    row.get("topic_conid"),
                    row.get("exchange"),
                    row.get("month"),
                    row.get("sectype"),
                    row.get("strike"),
                    row.get("has_yes"),
                    row.get("has_no"),
                    row.get("yes_conid"),
                    row.get("no_conid"),
                    row.get("yes_right"),
                    row.get("no_right"),
                    row.get("yes_description"),
                    row.get("no_description"),
                    row.get("yes_maturity_date"),
                    row.get("no_maturity_date"),
                    row.get("yes_trading_class"),
                    row.get("no_trading_class"),
                ]
            )

        workbook.save(output_path)
        print(f"Excel symbols export written to: {output_path}")


def chunks(items: list[str], size: int) -> list[list[str]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def month_variants() -> list[str]:
    now = datetime.now(tz=UTC)
    values = []
    for i in range(0, 6):
        y = now.year + ((now.month - 1 + i) // 12)
        m = ((now.month - 1 + i) % 12) + 1
        values.append(f"{y}{m:02d}")
    return values


def normalize_months(values: Any) -> list[str]:
    if not isinstance(values, list):
        return []
    out = []
    for value in values:
        text = str(value).strip().upper()
        if not text:
            continue
        if len(text) == 6 and text.isdigit():
            out.append(text)
            continue
        # Keep non-YYYYMM strings too, backend can handle variants.
        out.append(text)
    deduped: list[str] = []
    seen: set[str] = set()
    for item in out:
        if item not in seen:
            deduped.append(item)
            seen.add(item)
    return deduped


def load_symbols_from_excel(root: Path) -> list[str]:
    """Try reading any local xlsx for a symbol list."""
    if load_workbook is None:
        return []

    symbols: list[str] = []
    for xlsx_path in root.glob("*.xlsx"):
        try:
            workbook = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=1, max_col=5, values_only=True):
                for cell in row:
                    if not isinstance(cell, str):
                        continue
                    token = cell.strip().upper()
                    if not token or len(token) > 8:
                        continue
                    if not token.replace(".", "").replace("-", "").isalnum():
                        continue
                    symbols.append(token)
        except Exception:
            continue
    return symbols


def build_seed_symbols(user_symbols: list[str], cwd: Path) -> list[str]:
    defaults = [
        "FF",
        "USIP",
        "CPI",
        "GDP",
        "NQ",
        "ES",
        "RTY",
        "YM",
        "CL",
        "GC",
        "SI",
        "NG",
        "ZN",
        "ZB",
        "ZT",
        "VX",
    ]
    excel_symbols = load_symbols_from_excel(cwd)
    merged = defaults + user_symbols + excel_symbols
    deduped: list[str] = []
    seen: set[str] = set()
    for raw in merged:
        symbol = raw.strip().upper()
        if not symbol or symbol in seen:
            continue
        deduped.append(symbol)
        seen.add(symbol)
    return deduped


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Find prediction market symbols by running endpoint permutations and "
            "printing readable output for each fetch."
        )
    )
    parser.add_argument("--base-url", default="http://127.0.0.1:8000")
    parser.add_argument("--timeout", type=int, default=20)
    parser.add_argument("--symbol-limit", type=int, default=80)
    parser.add_argument("--chunk-size", type=int, default=12)
    parser.add_argument(
        "--symbols",
        default="",
        help="Comma-separated extra symbols to include in permutations.",
    )
    parser.add_argument(
        "--max-chain-topics",
        type=int,
        default=30,
        help="How many discovered topics to chain-probe for contract validation.",
    )
    parser.add_argument(
        "--output",
        default="prediction_market_discovery_report.json",
        help="Output JSON report path.",
    )
    parser.add_argument(
        "--excel-output",
        default="prediction_market_symbols.xlsx",
        help="Output Excel path for discovered prediction symbols.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    cwd = Path.cwd()
    user_symbols = [s.strip() for s in args.symbols.split(",") if s.strip()]
    symbols = build_seed_symbols(user_symbols, cwd)

    print("=" * 88)
    print("PREDICTION MARKET DISCOVERY RUNNER")
    print("=" * 88)
    print(f"Base URL: {args.base_url}")
    print(f"Candidate symbols ({len(symbols)}): {', '.join(symbols[:40])}")
    if len(symbols) > 40:
        print("... (truncated in console, full list included via report context)")

    runner = DiscoveryRunner(
        base_url=args.base_url,
        timeout=args.timeout,
        symbol_limit=args.symbol_limit,
        chunk_size=args.chunk_size,
        max_chain_topics=args.max_chain_topics,
    )

    runner.run_auth_checks()
    runner.run_topic_endpoint_permutations()
    runner.run_console_permutations(symbols)
    runner.run_search_permutations(symbols)
    runner.run_chain_permutations()
    runner.print_final_summary()
    runner.write_report(Path(args.output))
    runner.write_excel(Path(args.excel_output))

    return 0


if __name__ == "__main__":
    sys.exit(main())
